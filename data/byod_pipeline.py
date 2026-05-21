"""BYOD upload parsing and tier-1 battery feature extraction.

This module is deliberately dependency-light so the Render/lite service can
accept real CSV/TXT uploads without pandas. XLSX support is optional when
openpyxl is installed.
"""
from __future__ import annotations

import csv
import io
import math
import re
from collections import defaultdict
from dataclasses import dataclass
from statistics import mean, pstdev
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

FEATURE_NAMES = [
    "initial_capacity_ah",
    "early_coulombic_efficiency",
    "capacity_fade_rate_ah_per_cycle",
    "voltage_plateau_drop_v",
    "internal_resistance_ohm",
    "average_temperature_c",
    "estimated_c_rate",
    "estimated_cycle_at_80_soh",
    "voltage_variance_cc",
    "dqdv_peak_1_v",
    "dqdv_peak_2_v",
    "dqdv_peak_3_v",
    "sei_ce_slope",
    "early_degradation_acceleration",
    "final_capacity_fraction",
    "observed_cycle_count",
    "min_voltage_v",
    "max_voltage_v",
    "average_charge_current_a",
    "average_discharge_current_a",
    "initial_charge_capacity_ah",
    "initial_discharge_capacity_ah",
    "temperature_std_c",
    "rest_fraction",
    "data_completeness",
    "dqdv_peak_spread_v",
    "signal_confidence",
]

MIN_CAPACITY_SCALE = 0.01
MAX_FADE_FRACTION_PER_CYCLE = 0.02

CANONICAL_COLUMNS = [
    "time_seconds",
    "cycle_index",
    "step_index",
    "current_A",
    "voltage_V",
    "charge_capacity_ah",
    "discharge_capacity_ah",
    "capacity_ah",
    "temperature_C",
    "step_type",
]

VARIANTS: Dict[str, Sequence[str]] = {
    "time_seconds": [
        "time", "time_s", "time sec", "time(s)", "test_time_s", "test time seconds",
        "data_point_time", "step_time_s", "step time", "total_time", "total time",
        "elapsed_time", "elapsed time", "runtime", "record time", "date time",
        "relative time", "time/hms", "time/min", "time/hour", "time hours",
        "seconds", "s", "sec", "时间", "测试时间", "持续时间",
    ],
    "cycle_index": [
        "cycle", "cycle_index", "cycle index", "cycle_number", "cycle number",
        "cycle_no", "cycle no", "cycle_id", "cycle id", "loop", "loop index",
        "loop number", "cyc", "cycles", "循环", "循环号", "循环序号",
    ],
    "step_index": [
        "step", "step_index", "step index", "step_number", "step number",
        "step_no", "step no", "step_id", "procedure step", "record step",
        "工步", "步次", "工步号",
    ],
    "current_A": [
        "current", "current_a", "current(a)", "current [a]", "i", "i_a", "amps",
        "amp", "current ma", "current_ma", "current(mA)", "current [ma]",
        "test current", "current applied", "charge current", "discharge current",
        "电流", "电流(a)", "电流(ma)",
    ],
    "voltage_V": [
        "voltage", "voltage_v", "voltage(v)", "voltage [v]", "v", "ewe/v",
        "cell voltage", "terminal voltage", "potential", "potential/v",
        "u/v", "volt", "volts", "电压", "电压(v)",
    ],
    "capacity_ah": [
        "capacity", "capacity_ah", "capacity(ah)", "capacity [ah]", "q", "q_ah",
        "cap", "cap_ah", "total capacity", "step capacity", "specific capacity",
        "capacity mah", "capacity_mah", "capacity(mAh)", "q_discharge", "q charge",
        "容量", "容量(ah)", "容量(mah)",
    ],
    "charge_capacity_ah": [
        "charge capacity", "charge_capacity", "charge_capacity_ah", "charge cap",
        "chg cap", "chg_capacity", "chg. cap.", "charge cap ah", "charge_ah",
        "capacity charge", "q_charge", "q chg", "ch cap", "充电容量", "充电容量(ah)",
    ],
    "discharge_capacity_ah": [
        "discharge capacity", "discharge_capacity", "discharge_capacity_ah",
        "discharge cap", "dchg cap", "dchg_capacity", "dischg cap", "discharge_ah",
        "capacity discharge", "q_discharge", "q dchg", "dcap", "dch cap",
        "放电容量", "放电容量(ah)",
    ],
    "temperature_C": [
        "temperature", "temperature_c", "temperature(c)", "temperature [c]",
        "temp", "temp_c", "temp(c)", "ambient temperature", "cell temperature",
        "aux temp", "aux_temperature", "t1", "t2", "t/c", "温度", "温度(c)",
    ],
    "step_type": [
        "step type", "step_type", "mode", "state", "status", "command",
        "operation", "operation type", "test mode", "charge/discharge",
        "工步类型", "状态",
    ],
}

FORMAT_HINTS = {
    "neware": ["record index", "step index", "cycle index", "电流", "电压", "容量"],
    "arbin": ["data_point", "test_time", "step_time", "cycle_index"],
    "maccor": ["rec#", "cyc#", "step", "amps", "volts", "amp-hr"],
    "biologic": ["ewe/v", "i/ma", "time/s", "cycle number"],
    "basytec": ["line", "command", "u/v", "i/a", "ah"],
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\ufeff", "")
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    text = text.replace("[", "(").replace("]", ")")
    return text.strip()


def _is_specific_capacity_header(value: Any) -> bool:
    h = normalize_header(value)
    raw = str(value or "").lower()
    return "specific capacity" in h or "mah/g" in raw or "ma h/g" in raw or "ah/g" in raw or "capacity/g" in h


_VARIANT_LOOKUP = {
    normalize_header(alias): canonical
    for canonical, aliases in VARIANTS.items()
    for alias in aliases
}


def _column_score(header: str, canonical: str) -> float:
    h = normalize_header(header)
    if not h:
        return 0.0
    if canonical == "charge_capacity_ah" and any(token in h for token in ("discharge", "dchg", "dischg", "放电")):
        return 0.0
    if canonical == "discharge_capacity_ah" and ("charge" in h and not any(token in h for token in ("discharge", "dchg", "dischg"))):
        return 0.0
    if _VARIANT_LOOKUP.get(h) == canonical:
        return 1.0
    aliases = [normalize_header(a) for a in VARIANTS[canonical]]
    if any(h == a for a in aliases):
        return 1.0
    if any(a and a in h for a in aliases if len(a) >= 4):
        return 0.82
    tokens = set(re.findall(r"[a-z0-9]+", h))
    if canonical == "current_A" and ({"current"} & tokens or h in {"i", "i/a", "i/ma"}):
        return 0.72
    if canonical == "voltage_V" and ({"voltage", "volt", "potential"} & tokens or h in {"v", "u/v"}):
        return 0.72
    if canonical == "capacity_ah" and ({"capacity", "cap"} & tokens or h in {"q", "ah"}):
        return 0.64
    return 0.0


def detect_schema(headers: Sequence[Any]) -> Dict[str, Any]:
    mapping: Dict[str, str] = {}
    confidence: Dict[str, float] = {}
    used: set[str] = set()
    clean_headers = [str(h or "").strip() for h in headers]
    for canonical in CANONICAL_COLUMNS:
        best: Tuple[float, Optional[str]] = (0.0, None)
        for raw in clean_headers:
            if raw in used:
                continue
            score = _column_score(raw, canonical)
            if score > best[0]:
                best = (score, raw)
        if best[1] and best[0] >= 0.56:
            mapping[canonical] = best[1]
            confidence[canonical] = round(best[0], 3)
            used.add(best[1])

    haystack = " ".join(normalize_header(h) for h in clean_headers)
    detected_format = "generic"
    best_hits = 0
    for fmt, hints in FORMAT_HINTS.items():
        hits = sum(1 for hint in hints if normalize_header(hint) in haystack)
        if hits > best_hits:
            detected_format, best_hits = fmt, hits

    required = ["current_A", "voltage_V"]
    usable = sum(1 for c in required if c in mapping) == len(required)
    score = sum(confidence.values()) / max(1, len(CANONICAL_COLUMNS))
    capacity_headers = [
        mapping.get("capacity_ah"),
        mapping.get("charge_capacity_ah"),
        mapping.get("discharge_capacity_ah"),
    ]
    return {
        "mapping": mapping,
        "confidence": confidence,
        "format": detected_format,
        "score": round(score, 3),
        "usable": usable,
        "specific_capacity_detected": any(_is_specific_capacity_header(h) for h in capacity_headers if h),
        "missing": [c for c in CANONICAL_COLUMNS if c not in mapping],
        "unmatched_headers": [h for h in clean_headers if h and h not in used][:30],
    }


def _decode_text(content: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _read_delimited(content: bytes, max_rows: int) -> Tuple[List[str], List[Dict[str, str]], int]:
    text = _decode_text(content)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        counts = {d: sample.count(d) for d in [",", "\t", ";", "|"]}
        dialect = csv.excel_tab if max(counts, key=counts.get) == "\t" else csv.excel
        dialect.delimiter = max(counts, key=counts.get)
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        return [], [], 0
    header_idx = _find_header_row(rows[:30])
    headers = [str(x or "").strip() for x in rows[header_idx]]
    out: List[Dict[str, str]] = []
    for raw in rows[header_idx + 1: header_idx + 1 + max_rows]:
        if not any(str(x).strip() for x in raw):
            continue
        padded = list(raw) + [""] * max(0, len(headers) - len(raw))
        out.append({headers[i]: padded[i] for i in range(len(headers))})
    return headers, out, len(rows) - header_idx - 1


def _read_xlsx(content: bytes, max_rows: int) -> Tuple[List[str], List[Dict[str, str]], int]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency path
        raise ValueError("XLSX parsing requires openpyxl. Export CSV or install openpyxl.") from exc
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    blocks: List[Tuple[float, List[str], List[Sequence[Any]]]] = []
    raw_total = 0
    for ws in wb.worksheets:
        raw_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        if not raw_rows:
            continue
        header_idx = _find_header_row(raw_rows[:30])
        headers = [str(x or "").strip() for x in raw_rows[header_idx]]
        schema = detect_schema(headers)
        data_rows = raw_rows[header_idx + 1:]
        raw_total += len(data_rows)
        blocks.append((float(schema["score"]) + 0.25 * int(schema["usable"]), headers, data_rows))
    if not blocks:
        return [], [], 0
    blocks.sort(key=lambda item: item[0], reverse=True)
    headers = blocks[0][1]
    out: List[Dict[str, str]] = []
    for _score, sheet_headers, data_rows in blocks:
        if len(out) >= max_rows:
            break
        if out and not detect_schema(sheet_headers).get("usable"):
            continue
        for raw in data_rows:
            if len(out) >= max_rows:
                break
            if not any(str(x or "").strip() for x in raw):
                continue
            padded = list(raw) + [""] * max(0, len(sheet_headers) - len(raw))
            out.append({sheet_headers[i]: "" if padded[i] is None else str(padded[i]) for i in range(len(sheet_headers))})
    return headers, out, raw_total


def _find_header_row(rows: Sequence[Sequence[Any]]) -> int:
    def quick_score(row: Sequence[Any]) -> float:
        score = 0.0
        for cell in row:
            h = normalize_header(cell)
            if not h:
                continue
            if h in _VARIANT_LOOKUP:
                score += 1.0
            elif any(token in h for token in ("voltage", "current", "capacity", "cycle", "step", "time", "ewe/v", "i/ma")):
                score += 0.45
        return score

    candidates = sorted(
        ((quick_score(row), i, row) for i, row in enumerate(rows)),
        key=lambda item: item[0],
        reverse=True,
    )[:5]
    best_i, best_score = 0, -1.0
    for _quick, i, row in candidates:
        schema = detect_schema([str(x or "") for x in row])
        score = schema["score"] + 0.25 * int(schema["usable"])
        if score > best_score:
            best_i, best_score = i, score
    return best_i


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _unit_adjust(value: Optional[float], source_header: str, canonical: str) -> Optional[float]:
    if value is None:
        return None
    h = normalize_header(source_header)
    if canonical in {"current_A"} and ("ma" in h or "毫安" in h):
        return value / 1000.0
    if canonical.endswith("_ah") or canonical == "capacity_ah":
        if _is_specific_capacity_header(source_header):
            return value
        if "mah" in h or "毫安时" in h:
            return value / 1000.0
    if canonical == "time_seconds":
        if "min" in h:
            return value * 60.0
        if "hour" in h or h.endswith("/h"):
            return value * 3600.0
    return value


def _canonical_records(rows: Sequence[Dict[str, str]], mapping: Dict[str, str]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for raw in rows:
        rec: Dict[str, Any] = {}
        for canonical, header in mapping.items():
            value = raw.get(header)
            if canonical == "step_type":
                rec[canonical] = str(value or "").strip().lower()
            else:
                rec[canonical] = _unit_adjust(_to_float(value), header, canonical)
        current = rec.get("current_A")
        step_type = str(rec.get("step_type") or "")
        if not step_type and isinstance(current, (int, float)):
            if current > 1e-8:
                step_type = "charge"
            elif current < -1e-8:
                step_type = "discharge"
            else:
                step_type = "rest"
            rec["step_type"] = step_type
        out.append(rec)
    return out


def _finite(values: Iterable[Optional[float]]) -> List[float]:
    return [float(v) for v in values if isinstance(v, (int, float)) and math.isfinite(float(v))]


def _linear_fit(xs: Sequence[float], ys: Sequence[float]) -> Tuple[float, float]:
    if len(xs) < 2 or len(ys) < 2:
        return 0.0, ys[-1] if ys else 0.0
    mx, my = mean(xs), mean(ys)
    den = sum((x - mx) ** 2 for x in xs)
    if den <= 1e-12:
        return 0.0, my
    slope = sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / den
    return slope, my - slope * mx


def _cycle_summaries(records: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    inferred_cycle = 0
    last_cap: Optional[float] = None
    explicit_offset = 0
    explicit_segment_max = 0
    last_explicit_cycle: Optional[int] = None
    for rec in records:
        cyc = rec.get("cycle_index")
        if not isinstance(cyc, (int, float)) or not math.isfinite(float(cyc)):
            cap = rec.get("capacity_ah") or rec.get("discharge_capacity_ah") or rec.get("charge_capacity_ah")
            if last_cap is not None and isinstance(cap, (int, float)) and cap + 1e-6 < last_cap * 0.25:
                inferred_cycle += 1
            cyc = inferred_cycle
            last_cap = cap if isinstance(cap, (int, float)) else last_cap
        else:
            raw_cycle = int(round(float(cyc)))
            if last_explicit_cycle is not None and raw_cycle + 5 < last_explicit_cycle:
                explicit_offset += explicit_segment_max + 1
                explicit_segment_max = raw_cycle
            else:
                explicit_segment_max = max(explicit_segment_max, raw_cycle)
            last_explicit_cycle = raw_cycle
            cyc = raw_cycle + explicit_offset
        grouped[int(round(float(cyc)))].append(rec)
    summaries: List[Dict[str, Any]] = []
    for cyc in sorted(grouped):
        rows = grouped[cyc]
        dis_caps = _finite(r.get("discharge_capacity_ah") for r in rows)
        chg_caps = _finite(r.get("charge_capacity_ah") for r in rows)
        all_caps = _finite(r.get("capacity_ah") for r in rows)
        currents = _finite(r.get("current_A") for r in rows)
        voltages = _finite(r.get("voltage_V") for r in rows)
        temps = _finite(r.get("temperature_C") for r in rows)
        discharge_capacity = max(dis_caps) if dis_caps else (max(all_caps) if all_caps else None)
        charge_capacity = max(chg_caps) if chg_caps else None
        ce = discharge_capacity / charge_capacity if charge_capacity and discharge_capacity else None
        summaries.append({
            "cycle": cyc,
            "discharge_capacity_ah": discharge_capacity,
            "charge_capacity_ah": charge_capacity,
            "ce": ce,
            "current_abs_mean": mean(abs(i) for i in currents) if currents else None,
            "charge_current_mean": mean(i for i in currents if i > 1e-8) if any(i > 1e-8 for i in currents) else None,
            "discharge_current_mean": mean(abs(i) for i in currents if i < -1e-8) if any(i < -1e-8 for i in currents) else None,
            "voltage_mean": mean(voltages) if voltages else None,
            "voltage_min": min(voltages) if voltages else None,
            "voltage_max": max(voltages) if voltages else None,
            "voltage_var": pstdev(voltages) ** 2 if len(voltages) > 1 else None,
            "temperature_mean": mean(temps) if temps else None,
            "temperature_std": pstdev(temps) if len(temps) > 1 else None,
        })
    return summaries


def _first_discharge_trace(records: Sequence[Dict[str, Any]]) -> List[Tuple[float, float]]:
    by_cycle: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for rec in records:
        cyc = rec.get("cycle_index")
        if not isinstance(cyc, (int, float)):
            cyc = 0
        by_cycle[int(round(float(cyc)))].append(rec)
    for cyc in sorted(by_cycle):
        pts: List[Tuple[float, float]] = []
        for rec in by_cycle[cyc]:
            cap = rec.get("discharge_capacity_ah") or rec.get("capacity_ah")
            volt = rec.get("voltage_V")
            typ = str(rec.get("step_type") or "")
            current = rec.get("current_A")
            is_dis = "dis" in typ or (isinstance(current, (int, float)) and current < -1e-8)
            if is_dis and isinstance(cap, (int, float)) and isinstance(volt, (int, float)):
                pts.append((float(cap), float(volt)))
        if len(pts) >= 8:
            return pts
    return []


def _dqdv(trace: Sequence[Tuple[float, float]]) -> Dict[str, Any]:
    if len(trace) < 8:
        return {"points": [], "peaks": []}
    pts = sorted(trace, key=lambda x: x[1])
    compact: List[Tuple[float, float]] = []
    for cap, volt in pts:
        if not compact or abs(volt - compact[-1][1]) > 1e-4:
            compact.append((cap, volt))
    if len(compact) < 8:
        return {"points": [], "peaks": []}
    vals: List[Dict[str, float]] = []
    for i in range(2, len(compact) - 2):
        q0, v0 = compact[i - 2]
        q1, v1 = compact[i + 2]
        if abs(v1 - v0) < 1e-5:
            continue
        vals.append({"voltage": compact[i][1], "dqdv": (q1 - q0) / (v1 - v0)})
    if not vals:
        return {"points": [], "peaks": []}
    smoothed: List[Dict[str, float]] = []
    for i, row in enumerate(vals):
        win = vals[max(0, i - 2): min(len(vals), i + 3)]
        smoothed.append({"voltage": row["voltage"], "dqdv": mean(x["dqdv"] for x in win)})
    candidates: List[Tuple[float, float]] = []
    for i in range(1, len(smoothed) - 1):
        y = abs(smoothed[i]["dqdv"])
        if y >= abs(smoothed[i - 1]["dqdv"]) and y >= abs(smoothed[i + 1]["dqdv"]):
            candidates.append((y, smoothed[i]["voltage"]))
    peaks = [v for _, v in sorted(candidates, reverse=True)[:3]]
    peaks.sort()
    return {"points": smoothed[:: max(1, len(smoothed) // 160)], "peaks": peaks}


def _set_feature(values: List[float], mask: List[int], name: str, value: Optional[float]) -> None:
    idx = FEATURE_NAMES.index(name)
    if value is not None and math.isfinite(float(value)):
        values[idx] = float(value)
        mask[idx] = 1


def _observed_cycle_at_fraction(cycle_nums: Sequence[float], caps: Sequence[float], initial_cap: Optional[float], fraction: float = 0.80) -> Optional[float]:
    if not initial_cap or initial_cap <= 0 or not caps or not cycle_nums:
        return None
    threshold = initial_cap * fraction
    if caps[0] <= threshold:
        return float(cycle_nums[0])
    for i in range(1, min(len(caps), len(cycle_nums))):
        prev_cap, cap = caps[i - 1], caps[i]
        if cap <= threshold:
            prev_cycle, cycle = float(cycle_nums[i - 1]), float(cycle_nums[i])
            if abs(cap - prev_cap) <= 1e-12:
                return cycle
            frac = (threshold - prev_cap) / (cap - prev_cap)
            return prev_cycle + max(0.0, min(1.0, frac)) * (cycle - prev_cycle)
    return None


def _prediction_pack(features: Dict[str, float], mask_by_name: Dict[str, int], cycles: List[Dict[str, Any]]) -> Dict[str, Any]:
    observed = features.get("observed_cycle_count", 0.0)
    soh = features.get("final_capacity_fraction", 1.0)
    fade_rate = features.get("capacity_fade_rate_ah_per_cycle", 0.0)
    raw_init_cap = features.get("initial_capacity_ah", 1.0)
    capacity_scale_ok = isinstance(raw_init_cap, (int, float)) and math.isfinite(float(raw_init_cap)) and abs(float(raw_init_cap)) >= MIN_CAPACITY_SCALE
    init_cap = max(abs(float(raw_init_cap)) if isinstance(raw_init_cap, (int, float)) and math.isfinite(float(raw_init_cap)) else 1.0, MIN_CAPACITY_SCALE)
    fade_fraction_per_cycle = min(MAX_FADE_FRACTION_PER_CYCLE, max(0.0, -fade_rate / init_cap))
    if "estimated_cycle_at_80_soh" in features:
        cycle_80 = features["estimated_cycle_at_80_soh"]
    elif fade_fraction_per_cycle > 1e-9:
        cycle_80 = observed + max(0.0, soh - 0.80) / fade_fraction_per_cycle
    else:
        cycle_80 = None
    ce = features.get("early_coulombic_efficiency", 0.985)
    temp = features.get("average_temperature_c", 25.0)
    crate = features.get("estimated_c_rate", 1.0)
    voltage_max = features.get("max_voltage_v", 4.0)
    peak1 = features.get("dqdv_peak_1_v")
    peak3 = features.get("dqdv_peak_3_v", peak1 or 0.0)
    available = sum(mask_by_name.values())
    confidence = max(0.12, min(0.92, 0.18 + available / len(FEATURE_NAMES) * 0.66 - (0.16 if not capacity_scale_ok else 0.0)))
    electrolyte_degradation = max(0.0, min(1.0, (1.0 - ce) * 8.0 + max(0.0, temp - 30.0) * 0.006))
    plating_prob = max(0.0, min(1.0, (crate - 1.2) * 0.22 + max(0.0, 0.985 - ce) * 9.0 + max(0.0, 20.0 - temp) * 0.015))
    safe_crate = max(0.15, min(3.0, 2.5 - 1.4 * plating_prob - 0.5 * electrolyte_degradation))
    recovery_prob = max(0.0, min(1.0, 0.55 + (soh - 0.80) * 1.2 - max(0.0, features.get("early_degradation_acceleration", 0.0)) * 40.0))
    expected_gain = max(0.0, min(0.15, (1.0 - soh) * 0.18 * recovery_prob))
    if voltage_max < 3.7 and peak1 and 3.2 <= peak1 <= 3.55:
        chemistry = "LFP-like"
    elif voltage_max > 4.05 and peak3 and peak3 > 3.65:
        chemistry = "Na layered oxide / NFM-like"
    elif peak1 and peak1 > 3.55:
        chemistry = "NMC/NCA-like layered oxide"
    else:
        chemistry = "unknown layered/intercalation cell"
    formation_sei_quality = max(0.0, min(1.0, 0.35 + (ce - 0.90) * 5.0 - max(0.0, crate - 0.5) * 0.12))
    formation_life_index = max(0.0, min(1.0, 0.55 * formation_sei_quality + 0.45 * soh))
    formation_robustness = max(0.0, min(1.0, formation_life_index - max(0.0, temp - 35.0) * 0.006))
    def project_soh(target_cycle: float) -> float:
        extra = max(0.0, float(target_cycle) - observed)
        return max(0.0, min(1.05, soh - fade_fraction_per_cycle * extra))

    life_basis = cycle_80 if cycle_80 is not None and math.isfinite(cycle_80) else observed
    if life_basis < 500:
        life_bucket = "<500 cycles"
    elif life_basis < 1000:
        life_bucket = "500-1000 cycles"
    elif life_basis < 1500:
        life_bucket = "1000-1500 cycles"
    else:
        life_bucket = ">1500 cycles"
    voltage_var = features.get("voltage_variance_cc", 0.0)
    temp_std = features.get("temperature_std_c", 0.0)
    anomaly_score = max(0.0, min(1.0, (1.0 - confidence) * 0.45 + max(0.0, features.get("early_degradation_acceleration", 0.0)) * 70.0 + min(voltage_var, 0.05) * 3.0 + min(temp_std / 12.0, 0.25)))
    knee_cycle = None
    accel = features.get("early_degradation_acceleration")
    if accel is not None and accel < -1e-8 and fade_fraction_per_cycle > 0:
        knee_cycle = observed + min(1000.0, max(20.0, fade_fraction_per_cycle / abs(accel)))
    elif cycle_80 is not None and math.isfinite(cycle_80):
        knee_cycle = max(observed, cycle_80 * 0.72)
    bms_risk_proxy = max(0.0, min(1.0, max(0.0, temp - 35.0) * 0.012 + temp_std * 0.025 + max(0.0, crate - 1.5) * 0.18))
    chem_rank_score = max(0.0, min(1.0, 0.42 + 0.35 * soh + 0.12 * confidence - max(0.0, plating_prob - 0.25) * 0.18))
    return {
        "soh": round(soh, 5),
        "fade_fraction_per_cycle": round(fade_fraction_per_cycle, 8),
        "cycle_80_estimate": round(cycle_80, 1) if cycle_80 is not None and math.isfinite(cycle_80) else None,
        "confidence": round(confidence, 3),
        "model_outputs": {
            "M1_CathodeUDE": {
                "source": "production_rule_projection_from_tier1_features",
                "projected_soh_500": round(project_soh(500), 4),
                "projected_soh_1000": round(project_soh(1000), 4),
                "mechanism_hint": "SEI/electrolyte" if electrolyte_degradation > 0.28 else "capacity-fade dominated",
                "checkpoint_note": "Trained UDE inference is attached when the PyTorch runtime is available.",
            },
            "M2_SOH": {
                "source": "tier1_capacity_model",
                "soh": round(soh, 5),
                "confidence": round(confidence, 3),
            },
            "M3_CycleLife": {
                "source": "tier1_life_bucket",
                "cycle_life_bucket": life_bucket,
                "cycle_80_estimate": round(cycle_80, 1) if cycle_80 is not None and math.isfinite(cycle_80) else None,
            },
            "M4_FadeRate": {
                "source": "linear_early_cycle_fit",
                "fade_fraction_per_cycle": round(fade_fraction_per_cycle, 8),
                "capacity_fade_rate_ah_per_cycle": round(fade_rate, 9),
            },
            "M5_BMS_TGN": {
                "source": "single_cell_temperature_proxy",
                "pack_risk_proxy": round(bms_risk_proxy, 4),
                "requires_pack_geometry": True,
                "label_gate": True,
            },
            "M6_RUL": {
                "source": "cycle80_projection",
                "rul_cycles_to_80": round(max(0.0, cycle_80 - observed), 1) if cycle_80 is not None and math.isfinite(cycle_80) else None,
            },
            "M7_Anomaly": {
                "source": "feature_mask_residual_proxy",
                "anomaly_score": round(anomaly_score, 4),
                "low_confidence_driver": "missing_features" if available < 12 else "signal_shape",
            },
            "M8_Joint_SOH_RUL": {
                "source": "joint_rule_fusion",
                "soh": round(soh, 5),
                "cycle_80_estimate": round(cycle_80, 1) if cycle_80 is not None and math.isfinite(cycle_80) else None,
                "fade_fraction_per_cycle": round(fade_fraction_per_cycle, 8),
            },
            "M9_KneeDetect": {
                "source": "capacity_curvature_proxy",
                "knee_cycle_estimate": round(knee_cycle, 1) if knee_cycle is not None and math.isfinite(knee_cycle) else None,
            },
            "M10_ChemRank": {
                "source": "chemistry_family_rank_proxy",
                "predicted_family": chemistry,
                "rank_score": round(chem_rank_score, 4),
            },
            "M11_ElectrolyteHealth": {
                "source": "tier1_heuristic",
                "electrolyte_degradation": round(electrolyte_degradation, 4),
                "sodium_plating_probability": round(plating_prob, 4),
                "recommended_c_rate": round(safe_crate, 3),
            },
            "M12_Replenishability": {
                "source": "tier1_heuristic",
                "recovery_probability": round(recovery_prob, 4),
                "expected_recovery_fraction": round(expected_gain, 4),
                "research_preview": True,
            },
            "M13_ChemIdentifier": {
                "source": "dqdv_rule",
                "predicted_family": chemistry,
                "confidence": round(min(confidence, 0.72), 3),
            },
            "M14_FormationProtocol": {
                "source": "formation_rule",
                "life_index": round(formation_life_index, 4),
                "robustness_index": round(formation_robustness, 4),
                "sei_quality": round(formation_sei_quality, 4),
                "research_preview": True,
                "suggested_protocol": {
                    "formation_c_rate": round(max(0.03, min(0.35, 0.28 - 0.18 * (1.0 - formation_sei_quality))), 3),
                    "rest_time_hours": round(2.0 + 10.0 * (1.0 - formation_sei_quality), 2),
                    "note": "Derived from tier-1 features; validate against formation experiments before production use.",
                },
            },
        },
    }


def analyze_upload(filename: str, content: bytes, max_rows: int = 200_000) -> Dict[str, Any]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else "csv"
    if ext in {"xlsx", "xlsm"}:
        headers, rows, raw_count = _read_xlsx(content, max_rows=max_rows)
    elif ext in {"csv", "txt", "tsv", "mpt", "dat"}:
        headers, rows, raw_count = _read_delimited(content, max_rows=max_rows)
    else:
        raise ValueError(f"Unsupported upload type .{ext}; use CSV, TXT, TSV, MPT, or XLSX.")
    schema = detect_schema(headers)
    records = _canonical_records(rows, schema["mapping"])
    cycles = _cycle_summaries(records)
    dqdv = _dqdv(_first_discharge_trace(records))

    values = [0.0] * len(FEATURE_NAMES)
    mask = [0] * len(FEATURE_NAMES)
    caps = [c["discharge_capacity_ah"] for c in cycles if isinstance(c.get("discharge_capacity_ah"), (int, float))]
    cycle_nums = [float(c["cycle"]) for c in cycles if isinstance(c.get("discharge_capacity_ah"), (int, float))]
    initial_cap = caps[0] if caps else None
    final_fraction = caps[-1] / initial_cap if initial_cap and caps else None
    fit_n = min(50, len(caps))
    fade_slope = _linear_fit(cycle_nums[:fit_n], caps[:fit_n])[0] if fit_n >= 2 else None
    ce_vals = [c["ce"] for c in cycles[:10] if isinstance(c.get("ce"), (int, float)) and c["ce"] > 0]
    ce_early = mean(ce_vals) if ce_vals else None
    ce_slope = _linear_fit([float(i) for i in range(len(ce_vals))], ce_vals)[0] if len(ce_vals) >= 3 else None
    accel = None
    if len(caps) >= 5:
        second = [caps[i] - 2 * caps[i - 1] + caps[i - 2] for i in range(2, min(len(caps), 20))]
        accel = mean(second) if second else None
    temps = _finite(r.get("temperature_C") for r in records)
    currents = _finite(r.get("current_A") for r in records)
    voltages = _finite(r.get("voltage_V") for r in records)
    charge_curr = [i for i in currents if i > 1e-8]
    discharge_curr = [abs(i) for i in currents if i < -1e-8]
    rest_fraction = sum(1 for i in currents if abs(i) <= 1e-8) / max(1, len(currents)) if currents else None
    c_rate = (mean([abs(i) for i in currents]) / initial_cap) if currents and initial_cap else None
    if schema.get("specific_capacity_detected"):
        c_rate = None
    cycle_80 = _observed_cycle_at_fraction(cycle_nums, caps, initial_cap, 0.80)
    if cycle_80 is None and initial_cap and fade_slope and fade_slope < 0 and caps and cycle_nums:
        threshold = 0.8 * initial_cap
        if caps[-1] > threshold:
            projected = float(cycle_nums[-1]) + (threshold - caps[-1]) / fade_slope
            if projected >= float(cycle_nums[-1]):
                cycle_80 = projected
    plateau_drop = None
    if len(cycles) >= 2:
        first_v = cycles[0].get("voltage_mean")
        target = cycles[min(len(cycles) - 1, 49)].get("voltage_mean")
        if isinstance(first_v, (int, float)) and isinstance(target, (int, float)):
            plateau_drop = first_v - target
    r_int = None
    if len(records) >= 3:
        best = 0.0
        for a, b in zip(records, records[1:]):
            ia, ib = a.get("current_A"), b.get("current_A")
            va, vb = a.get("voltage_V"), b.get("voltage_V")
            if all(isinstance(x, (int, float)) for x in [ia, ib, va, vb]) and abs(ib - ia) > 1e-6:
                best = max(best, abs((vb - va) / (ib - ia)))
        r_int = best or None

    _set_feature(values, mask, "initial_capacity_ah", initial_cap)
    _set_feature(values, mask, "early_coulombic_efficiency", ce_early)
    _set_feature(values, mask, "capacity_fade_rate_ah_per_cycle", fade_slope)
    _set_feature(values, mask, "voltage_plateau_drop_v", plateau_drop)
    _set_feature(values, mask, "internal_resistance_ohm", r_int)
    _set_feature(values, mask, "average_temperature_c", mean(temps) if temps else None)
    _set_feature(values, mask, "estimated_c_rate", c_rate)
    _set_feature(values, mask, "estimated_cycle_at_80_soh", cycle_80)
    _set_feature(values, mask, "voltage_variance_cc", pstdev(voltages) ** 2 if len(voltages) > 1 else None)
    for i, peak in enumerate(dqdv["peaks"][:3], start=1):
        _set_feature(values, mask, f"dqdv_peak_{i}_v", peak)
    _set_feature(values, mask, "sei_ce_slope", ce_slope)
    _set_feature(values, mask, "early_degradation_acceleration", accel)
    _set_feature(values, mask, "final_capacity_fraction", final_fraction)
    _set_feature(values, mask, "observed_cycle_count", float(len(caps)) if caps else None)
    _set_feature(values, mask, "min_voltage_v", min(voltages) if voltages else None)
    _set_feature(values, mask, "max_voltage_v", max(voltages) if voltages else None)
    _set_feature(values, mask, "average_charge_current_a", mean(charge_curr) if charge_curr else None)
    _set_feature(values, mask, "average_discharge_current_a", mean(discharge_curr) if discharge_curr else None)
    _set_feature(values, mask, "initial_charge_capacity_ah", cycles[0].get("charge_capacity_ah") if cycles else None)
    _set_feature(values, mask, "initial_discharge_capacity_ah", cycles[0].get("discharge_capacity_ah") if cycles else None)
    _set_feature(values, mask, "temperature_std_c", pstdev(temps) if len(temps) > 1 else None)
    _set_feature(values, mask, "rest_fraction", rest_fraction)
    _set_feature(values, mask, "data_completeness", sum(mask) / len(mask))
    if len(dqdv["peaks"]) >= 2:
        _set_feature(values, mask, "dqdv_peak_spread_v", max(dqdv["peaks"]) - min(dqdv["peaks"]))
    _set_feature(values, mask, "signal_confidence", max(0.0, min(1.0, 0.18 + sum(mask) / len(mask) * 0.72)))

    feature_dict = {name: values[i] for i, name in enumerate(FEATURE_NAMES) if mask[i]}
    mask_by_name = {name: mask[i] for i, name in enumerate(FEATURE_NAMES)}
    cycle_preview = [
        {
            "cycle": c["cycle"],
            "discharge_capacity_ah": c.get("discharge_capacity_ah"),
            "charge_capacity_ah": c.get("charge_capacity_ah"),
            "ce": c.get("ce"),
        }
        for c in cycles[:160]
    ]
    return {
        "filename": filename,
        "rows_read": len(rows),
        "rows_available": raw_count,
        "truncated": raw_count > len(rows),
        "schema": schema,
        "features": feature_dict,
        "feature_vector": values,
        "feature_mask": mask,
        "feature_names": FEATURE_NAMES,
        "feature_availability": mask_by_name,
        "cycle_summary": cycle_preview,
        "dqdv": dqdv,
        "predictions": _prediction_pack(feature_dict, mask_by_name, cycles),
        "warnings": _warnings(schema, feature_dict, mask, raw_count, len(rows)),
    }


def _warnings(schema: Dict[str, Any], features: Dict[str, float], mask: Sequence[int], raw_count: int, rows_read: int) -> List[str]:
    out: List[str] = []
    if not schema.get("usable"):
        out.append("Schema confidence is low. Confirm current and voltage mappings before trusting diagnostics.")
    if schema.get("specific_capacity_detected"):
        out.append("Specific capacity units were detected. SOH trends are usable, but absolute Ah, C-rate, and pack/BMS outputs need active-material mass before production use.")
    if raw_count > rows_read:
        out.append(f"File was truncated to {rows_read} rows for upload analysis. Use local/offline deployment for full multi-GB cycler files.")
    init_cap = features.get("initial_capacity_ah")
    if isinstance(init_cap, (int, float)) and math.isfinite(float(init_cap)) and abs(float(init_cap)) < MIN_CAPACITY_SCALE:
        out.append("Initial capacity scale is below 0.01 Ah or Ah-equivalent. Capacity-derived projections were capped to prevent nonsense extrapolation.")
    if "early_coulombic_efficiency" not in features:
        out.append("Charge/discharge capacity columns were not both detected, so CE and formation quality are lower confidence.")
    if "dqdv_peak_1_v" not in features:
        out.append("dQ/dV peaks could not be extracted; provide discharge voltage-capacity traces for chemistry identification.")
    if sum(mask) < 10:
        out.append("Fewer than 10 tier-1 features were available. Treat ML-style outputs as low-confidence triage only.")
    return out
